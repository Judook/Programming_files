try :
    import math
    import random
    import keyboard
    import openpyxl
    
except :
    print("Some libraries for this application to work are not installed. Please install those libraries and run this application again.")
    print("Required libraries are :\n1. math\n2. random\n3. keyboard\n4. openpyxl")
    input("Press Enter to continue...")
    exit()

def split_marks(marks, out_of, splits) :
    splitted = []
    splitted_sum = 0
    for item in splits :
        splitted.append(math.floor((marks / out_of) * item))

    for item in splitted :
        splitted_sum += item

    if splitted_sum == marks :
        return splitted
    else :
        difference = marks - splitted_sum

        for i in range(difference) :
            while True :
                random_number = random.randint(0, len(splitted) - 1)
                if splitted[random_number] <= splits[random_number] :
                    splitted[random_number] += 1
                    break

        return splitted
    
while True :    
    try :
        mark_list = openpyxl.load_workbook(str(input("Enter the worksheet file Location : ")))
        break
    except :
        print("Enter a valid location!")

splits = int(input("Enter the number of parts of marks : "))
split_lst = []
max_mark = 0

for i in range(splits) :
    split_lst.append(int(input("Enter the maximum mark of part " + str(i+1) + " : ")))
    max_mark += split_lst[len(split_lst) - 1]

print("Total maximum marks is", max_mark)

mark_list = mark_list.active
lst = []
out = []

for row in range(0, mark_list.max_row):
    for col in mark_list.iter_cols(1, mark_list.max_column):
        lst.append(col[row].value)

for item in lst :
    out.append(split_marks(item, max_mark, split_lst))

print("Open an excel worksheet, place the active cell at A1 and press F9 to start printing the output")
print("Please do not press any key while printing the output")
keyboard.wait("F9")

keyboard.write("Marks")
keyboard.press_and_release("right")

for item in split_lst :
    keyboard.write("O.O." + str(item))
    keyboard.press_and_release("right")
    
keyboard.press_and_release("down")
keyboard.press_and_release("home")

for i in range(len(out)) :
    keyboard.write(str(lst[i]))
    keyboard.press_and_release("right")

    for j in range(len(out[i])) :
        keyboard.write(str(out[i][j]))
        keyboard.press_and_release("right")

    keyboard.press_and_release("down")
    keyboard.press_and_release("home")
