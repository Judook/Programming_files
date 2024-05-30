try :
    import math
    import random
    import keyboard
    import openpyxl
    import tkinter
    
except :
    print("Some libraries for this application to work are not installed. Please install those libraries and run this application again.")
    print("Required libraries are :\n1. math\n2. random\n3. keyboard\n4. openpyxl\n5. tkinter")
    input("Press Enter to continue...")
    exit()

def split_marks(marks, out_of, splits) :
    splitted = []
    splitted_sum = 0
    for item in splits :
        splitted.append(math.floor((marks / out_of) * item))

    for item in splitted :
        splitted_sum += item

    if splitted_sum == marks :
        return splitted
    else :
        difference = marks - splitted_sum

        for i in range(difference) :
            while True :
                random_number = random.randint(0, len(splitted) - 1)
                if splitted[random_number] <= splits[random_number] :
                    splitted[random_number] += 1
                    break

        return splitted
    
def set_location() :
    try :
        global mark_list
        global set_location_bool
        location = tkinter.filedialog.askopenfilename(filetypes = [("Excel files", "*.xlsx")])
        mark_list = openpyxl.load_workbook(location)
        location_error_label.config(text = "")
        location_info_label.config(text = "File : " + location)
        
        set_location_bool = True
    except :
        location_error_label.config(text = "Please provide a valid location!", foreground = "red")
        set_location_bool = False

def set_max_mark_list() :
    try :
        global split_lst
        global set_max_mark_list_bool
        split_lst = []
        max_mark_list = list(mark_splits_text.get())
        mark = ""
        for item in max_mark_list :
            if item != "," :
                mark += item
            else :
                split_lst.append(int(mark))
                mark = ""
        split_lst.append(int(mark))
        mark_splits_error_label.config(text = "")
        set_max_mark_list_bool = True
    except :
        mark_splits_error_label.config(text = "Enter the list of maximum marks correctly!", foreground = "red")
        set_max_mark_list_bool = False

def set_max_marks() :
        global max_mark
        max_mark = 0
        for item in split_lst :
            max_mark += item
        max_mark_label.config(text = "Total maximum marks is " + str(max_mark))

def start_print() :
    if set_max_mark_list_bool and set_location_bool:
        keyboard.wait("F9")

        keyboard.write("Marks")
        keyboard.press_and_release("right")

        for item in split_lst :
            keyboard.write("O.O." + str(item))
            keyboard.press_and_release("right")#
            
        keyboard.press_and_release("down")
        keyboard.press_and_release("home")

        for i in range(len(out)) :
            keyboard.write(str(lst[i]))
            keyboard.press_and_release("right")

            for j in range(len(out[i])) :
                keyboard.write(str(out[i][j]))
                keyboard.press_and_release("right")

            keyboard.press_and_release("down")
            keyboard.press_and_release("home")   

def set_everything() :
    set_max_mark_list()
    set_max_marks()

    global mark_list
    global lst
    global out

    if set_max_mark_list_bool and set_location_bool :
        mark_list = mark_list.active
        lst = []
        out = []

        for row in range(0, mark_list.max_row):
            for col in mark_list.iter_cols(1, mark_list.max_column):
                lst.append(col[row].value)

        for item in lst :
            out.append(split_marks(item, max_mark, split_lst))

        instruction_label.config(text = "Press Start, Open an excel worksheet, place the active cell at first column and press F9 to start printing the output")
        instruction_label_2.config(text = "Please do not press any key while printing the output")

        start_button.pack(padx = 10, pady = 10)

window = tkinter.Tk()
window.config(background = "#1a1a1a")
window.geometry("810x640")
window.title("Mark Splitter")
window.resizable(width = False, height = False)

big_title = tkinter.Label(window, text = "Mark Splitter", font = ("Arial", 36), foreground = "green", background = "#1a1a1a")

location_label = tkinter.Label(window, text = "Browse the file location of the .xlsx file", font = ("Arial", 15), background = "#1a1a1a", foreground = "white")
location_info_label = tkinter.Label(window, text = "", font = ("Arial", 12), background = "#1a1a1a", foreground = "white")
location_button = tkinter.Button(window, borderwidth = 0, text = "Browse", command = set_location, width = 20, height = 2, background = "green", foreground = "white", activebackground = "#3b3b3b", activeforeground = "white")
location_error_label = tkinter.Label(window, text = "", font = ("Arial", 12), background = "#1a1a1a")

mark_splits_label = tkinter.Label(window, text = "Enter the list of maximum marks for different parts seperated by a comma(,)", font = ("Arial", 15), background = "#1a1a1a", foreground = "white")
mark_splits_text = tkinter.Entry(window, width = 450, font = ("Arial", 15), borderwidth = 0, background = "#363636", foreground = "white")
mark_splits_error_label = tkinter.Label(window, text = "", font = ("Arial", 12), background = "#1a1a1a")

max_mark_label = tkinter.Label(window, text = "", font = ("Arial", 15), background = "#1a1a1a", foreground = "white")

set_button = tkinter.Button(window, borderwidth = 0, text = "Set", command = set_everything, width = 20, height = 2, background = "green", foreground = "white", activebackground = "#3b3b3b", activeforeground = "white")

instruction_label = tkinter.Label(window, text = "", font = ("Arial", 12), width = 450, background = "#1a1a1a", foreground = "white")
instruction_label_2 = tkinter.Label(window, text = "", font = ("Arial", 12), width = 450, background = "#1a1a1a", foreground = "white")

start_button = tkinter.Button(window, text = "Start", borderwidth = 0, width = 20, height = 2, command = start_print, background = "green", foreground = "white", activebackground = "#3b3b3b", activeforeground = "white")

big_title.pack(padx = 10, pady = 10)

location_label.pack(padx = 10, pady = 10)
location_info_label.pack(padx = 10, pady = 10)
location_button.pack(padx = 10, pady = 10)
location_error_label.pack()

mark_splits_label.pack(padx = 10, pady = 10)
mark_splits_text.pack(padx = 10, pady = 10)
mark_splits_error_label.pack()

max_mark_label.pack(padx = 10, pady = 10)

set_button.pack(padx = 10, pady = 10)

instruction_label.pack(padx = 10, pady = 10)
instruction_label_2.pack(padx = 10, pady = 10)

window.mainloop()
