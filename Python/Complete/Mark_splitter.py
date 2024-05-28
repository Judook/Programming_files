try :
    import math
    import random
    import keyboard
    import openpyxl
    
except :
    print("Some libraries for this application to work are not installed. Please install those libraries and run this application again.")
    print("Required libraries are :\n1. math\n2. random\n3. keyboard\n4. openpyxl")
    input("Press Enter to continue...")

def split_marks(marks) :
    splitted = []
    splitted.append(math.floor((marks/50) * 25))
    splitted.append(math.floor((marks/50) * 15))
    splitted.append(math.floor((marks/50) * 10))

    if splitted[0] + splitted[1] + splitted[2] == marks :
        return splitted
    else :
        diff = marks - (splitted[0] + splitted[1] + splitted[2])
        for i in range(diff) :
            if splitted[0] / 25 < splitted[1] / 15 and splitted[0] / 25 < splitted[2] / 10 :
                splitted[0] += 1

            elif splitted[1] / 15 < splitted[0] / 25 and splitted[1] / 15 < splitted[2] / 10 :
                splitted[1] += 1

            elif splitted[2] / 10 < splitted[0] / 25 and splitted[2] / 10 < splitted[1] / 15 :
                splitted[2] += 1

            else :
                splitted[random.randint(0, 2)] += 1

        return splitted

while True :    
    try :
        mark_list = openpyxl.load_workbook(str(input("Enter the worksheet file Location : ")))
        break
    except :
        print("Enter a valid location!")

mark_list = mark_list.active
lst = []
out = []

for row in range(0, mark_list.max_row):
    for col in mark_list.iter_cols(1, mark_list.max_column):
        lst.append(col[row].value)

for item in lst :
    out.append(split_marks(item))

print("Open an excel worksheet, place the active cell at A1 and press F9 to start printing the output")
print("Please do not press any key while printing the output")
keyboard.wait("F9")

keyboard.write("Marks")
keyboard.press_and_release("right")
keyboard.write("O.O.25")
keyboard.press_and_release("right")
keyboard.write("O.O.15")
keyboard.press_and_release("right")
keyboard.write("O.O.10")
keyboard.press_and_release("down")
keyboard.press_and_release("home")

for i in range(len(out)) :
    keyboard.write(str(lst[i]))
    keyboard.press_and_release("right")

    for j in range(len(out[i])) :
        keyboard.write(str(out[i][j]))
        keyboard.press_and_release("right")

    keyboard.press_and_release("down")
    keyboard.press_and_release("home")
